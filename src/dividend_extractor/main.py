import yfinance as yf
import pandas as pd
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle
import os

# GLOBAL VARIABLES AND SETTINGS
date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')

def save_to_excel(records):
    '''
    Saves the records to a new Excel file.

    :param records: list containing funds and dividend information (Date, Ticker, Dividends, Shares)
    :return: excel file with information of the records variable.
    '''

    # Creates a new Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dividends"

    # Column headers
    headers = ['Date', 'Ticker', 'Dividend', 'Shares']
    ws.append(headers)

    # Adds the records to the worksheet
    for row_index, dividend in enumerate(records, start=2):
        row = [
            dividend["date"],
            dividend["ticker"],
            dividend["dividends"],
            dividend["shares"]
        ]

        for col_index, item in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            if isinstance(item, datetime):
                cell.value = item
                cell.style = date_style
            else:
                cell.value = item

            # Adjusts the column width based on content
            column_letter = get_column_letter(col_index)
            size = len(str(cell.value))
            ws.column_dimensions[column_letter].width = size + 2

    filename = "Dividends.xlsx"
    wb.save(filename)
    os.startfile(filename)

def extract_date(date_string):
    '''
    Extracts the day, month, and year from a string in the format DD/MM/YYYY.

    :param date_string (str): Date in the format DD/MM/YYYY.
    :return: A tuple containing day, month, and year.

    >>>extract_date("01/03/2025")
    (1, 3, 2025)
    '''

    try:
        # Splits the string to get day, month, and year
        day, month, year = map(int, date_string.split('/'))

        # List of days in each month (index 0 is january, index 11 is december)
        days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

        # Validation for the values of day, month, and year
        if not (1 <= month <= 12):
            raise ValueError("Invalid month. It must be between 1 and 12.")

        # Adjust february for leap year
        if month == 2:
            if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
                days_in_month[1] = 29

        # Validate the day according to the month
        if not (1 <= day <= days_in_month[month - 1]):
            raise ValueError(f"Invalid day for month {month}. This month has {days_in_month[month - 1]} days.")

        return day, month, year

    except ValueError as e:
        print(f"Error: {e}")
        return None

def get_dividends(ticker: str):
    '''
    Gets the dividends paid by a stock using the yfinance library.

    :param ticker (str): The REIT (FII) code (e.g., 'BTLG11.SA').
    :return dataFrame: A dataframe containing the dates and dividend values.

    >>>get_dividends("RECR11.SA")
    2022-04-08    1.365384
    2022-05-09    1.502400
    2022-06-08    1.720500
    2022-07-08    1.150000
    2022-08-08    1.020500
    2022-09-09    0.800000
    '''

    try:
        # Downloads the historical data for the stock
        fii = yf.Ticker(ticker)

        # Gets the dividends
        dividends = fii.dividends

        # If the dividends dataframe is empty, returns none
        if dividends.empty:
            return None
        
        # Ensure that dividends index is a DatetimeIndex before attempting to localize the timezone
        if isinstance(dividends.index, pd.RangeIndex):
            return None  # Return none if there's no datetime index

        # Remove the timezone if it exists
        if dividends.index.tz is not None:
            dividends.index = dividends.index.tz_localize(None)

        if dividends.empty:
            return None

        # Returns dividends as a DataFrame
        return dividends

    except Exception as e:
        print(f"Error getting dividends for {ticker}: {e}")
        return None

if __name__ == "__main__":

    # Defines the start date for dividend search
    start_date = "01/03/2025"
    start_date = extract_date(start_date)

    # Defines the end date for dividend search
    end_date = "31/03/2025"
    end_date = extract_date(end_date)

    if start_date and end_date:
        # Defines the month limits for the search
        start_of_month = datetime(start_date[2], start_date[1], start_date[0])
        end_of_month = datetime(end_date[2], end_date[1], end_date[0])

        # API scope configuration to read from GoogleSheets
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

        # Authentication with service account credentials
        credentials = ServiceAccountCredentials.from_json_keyfile_name("src/dividend_extractor/credentials/credentials.json", scope)
        client = gspread.authorize(credentials)

        # Searching for the spreadsheet by name
        spreadsheet = client.open("Planejamento 2025")

        # Selecting the 'Portfolio' tab for reading
        sheet = spreadsheet.worksheet("Carteira")
        rows = sheet.get_all_values()

        # Asset list
        assets = []

        for row in rows[1:]:
            # Checks if a fund is registered in the sheet
            if row[0]:
                if int(row[4]) > 0: # Proceeds with extraction only if shares are available

                    # Adds the '.SA' prefix to refer to the São Paulo Stock Exchange - Brazil
                    ticker = row[0] + ".SA"

                    # Gets all dividends for the FII
                    dividend_list = get_dividends(ticker)
                    
                    if dividend_list is not None:
                        # Filters dividends based on the given period
                        dividend_period = dividend_list[(dividend_list.index >= start_of_month) & (dividend_list.index <= end_of_month)]

                    # If there are dividends, gets the dividend distribution date
                    if not dividend_period.empty:
                        distribution_date = dividend_period.index

                    # If there are dividends, extracts and adds them to the list
                        record = {
                            "date": distribution_date[0].strftime("%d/%m/%Y"),
                            "ticker": ticker[:-3],
                            "dividends": float(dividend_period.sum()) * int(row[4]),
                            "shares": int(row[4])
                        }

                        # Adds the record to the asset list
                        assets.append(record)
            else:
                break  # Exits the loop when an empty row is found in the first column

        # Sorts the list by dividend payment date
        sorted_assets = sorted(
            assets,
            key=lambda x: datetime.strptime(x["date"], "%d/%m/%Y")
        )
        save_to_excel(sorted_assets)
